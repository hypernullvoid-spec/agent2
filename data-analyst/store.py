"""
Storage layer for LARGE files.

The idea: convert every upload to Parquet ONCE, then never load the rows into
Python again. All reading happens through DuckDB, which queries the file on
disk. Memory stays flat whether the table has 1,000 rows or 50,000,000.

Use this instead of loader.py when files are big or when many users upload
at the same time.

    pip install duckdb
"""

import csv
import os
import re

import duckdb

STORE_DIR = "store"
os.makedirs(STORE_DIR, exist_ok=True)


def _squash(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


# ----------------------------------------------------------------------
# INGEST - runs once per upload
# ----------------------------------------------------------------------

def ingest(path, file_id):
    """
    Turn an upload into one Parquet file per table.
    Streams throughout, so a 500 MB file uses very little memory.

    Returns {table_name: parquet_path}.
    """
    if path.lower().endswith((".xlsx", ".xls")):
        csv_paths = _excel_to_csvs(path, file_id)
    else:
        csv_paths = {"data": path}

    tables = {}
    for name, csv_path in csv_paths.items():
        parquet_path = os.path.join(STORE_DIR, f"{file_id}_{name}.parquet")

        # DuckDB reads the CSV and writes Parquet without loading it into Python.
        # Parquet is columnar, so reading 2 of 30 columns later touches only those 2.
        duckdb.sql(f"""
            COPY (SELECT * FROM read_csv_auto('{csv_path}', ignore_errors=true))
            TO '{parquet_path}' (FORMAT PARQUET, COMPRESSION ZSTD)
        """)
        tables[name] = parquet_path

    return tables


def _excel_to_csvs(path, file_id):
    """Stream each Excel sheet straight out to a CSV file. Never holds the sheet in RAM."""
    import openpyxl

    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    written = {}

    try:
        for name in book.sheetnames:
            safe = re.sub(r"[^A-Za-z0-9_]+", "_", name).strip("_") or "sheet"
            out_path = os.path.join(STORE_DIR, f"{file_id}_{safe}.csv")

            with open(out_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                headers, carry = None, {}

                for raw in book[name].iter_rows(values_only=True):
                    cells = [_squash(c) for c in raw]

                    if headers is None:
                        if sum(1 for c in cells if c) >= max(2, len(cells) * 0.6):
                            headers = [c or f"column_{i}" for i, c in enumerate(cells)]
                            writer.writerow(headers)
                        continue

                    for i, col in enumerate(headers):        # fill merged cells downward
                        if i < len(cells):
                            if cells[i] == "":
                                cells[i] = carry.get(col, "")
                            else:
                                carry[col] = cells[i]

                    if any(cells):
                        writer.writerow(cells[:len(headers)])

            if headers:
                written[safe] = out_path
            else:
                os.remove(out_path)
    finally:
        book.close()

    return written


# ----------------------------------------------------------------------
# READ - all of these are constant memory, whatever the table size
# ----------------------------------------------------------------------

def schema(parquet_path):
    """Column names, types and row count. Reads metadata only - no data."""
    columns = duckdb.sql(f"DESCRIBE SELECT * FROM '{parquet_path}'").fetchall()
    rows = duckdb.sql(f"SELECT count(*) FROM '{parquet_path}'").fetchone()[0]
    return {"rows": rows, "columns": [{"name": c[0], "type": c[1]} for c in columns]}


def column_weights(parquet_path):
    """
    How heavy is each column? Measured with SQL, so nothing is loaded.
    Returns [(column, percent, rough_tokens)] heaviest first.
    """
    info = schema(parquet_path)
    names = [c["name"] for c in info["columns"]]
    if not names:
        return []

    sums = ", ".join(f'sum(length(coalesce(cast("{n}" as varchar), \'\'))) AS "{n}"' for n in names)
    values = duckdb.sql(f"SELECT {sums} FROM '{parquet_path}'").fetchone()

    sizes = dict(zip(names, (v or 0 for v in values)))
    total = sum(sizes.values()) or 1

    return sorted(
        ((n, size / total * 100, int(size / 3.5)) for n, size in sizes.items()),
        key=lambda item: -item[1],
    )


def sample(parquet_path, n=5):
    """A few example rows, for showing the model what the data looks like."""
    return duckdb.sql(f"SELECT * FROM '{parquet_path}' LIMIT {int(n)}").fetchall()


def to_csv_text(parquet_path, columns, where=None, limit=None):
    """
    Build the CSV text to send to the model - only the columns asked for.

    Because Parquet is columnar, selecting 3 of 30 columns reads only those 3
    off the disk. This is the whole reason for using Parquet.
    """
    picked = ", ".join(f'"{c}"' for c in columns)
    query = f"SELECT {picked} FROM '{parquet_path}'"
    if where:
        query += f" WHERE {where}"
    if limit:
        query += f" LIMIT {int(limit)}"

    result = duckdb.sql(query)
    lines = [",".join(columns)]
    for row in result.fetchall():
        lines.append(",".join(_csv_cell(v) for v in row))
    return "\n".join(lines)


def _csv_cell(value):
    text = "" if value is None else str(value)
    if any(ch in text for ch in ',"\n'):
        return '"' + text.replace('"', '""') + '"'
    return text


def sample_rows(parquet_path, n=800):
    """
    A few rows as plain dicts, so loader.classify_columns() can inspect them.

    Reads only n rows off disk, so this is cheap on a 900,000-row table.
    """
    result = duckdb.sql(f"SELECT * FROM '{parquet_path}' LIMIT {int(n)}")
    names = [d[0] for d in result.description]
    return [{k: ("" if v is None else str(v)) for k, v in zip(names, row)}
            for row in result.fetchall()]


def rollup_rows(parquet_path, group_by, measures, min_rows=1, limit=5000):
    """
    Group a huge table down to a small one, entirely in SQL.

    Returns plain dicts shaped exactly like loader.rollup(), so analyst.ask()
    cannot tell the difference. 900,000 rows in, a few hundred out, and the
    rows never enter Python.
    """
    groups = ", ".join(f'"{c}"' for c in group_by)
    aggs = ", ".join(f'round(avg(try_cast("{c}" AS DOUBLE)), 2) AS "avg_{c}"'
                     for c in measures)
    select = f"{groups}, count(*) AS n_rows" + (f", {aggs}" if aggs else "")

    result = duckdb.sql(f"""
        SELECT {select}
        FROM '{parquet_path}'
        WHERE {' AND '.join(f'"{c}" IS NOT NULL' for c in group_by)}
        GROUP BY {groups}
        HAVING count(*) >= {int(min_rows)}
        ORDER BY {groups}
        LIMIT {int(limit)}
    """)

    names = [d[0] for d in result.description]
    return [{k: ("" if v is None else v) for k, v in zip(names, row)}
            for row in result.fetchall()]


def rollup(parquet_path, group_by, measures):
    """
    Shrink a huge table by grouping it.
    8 million transaction rows -> 4,000 summary rows, done entirely in SQL.

        rollup(path, ["region", "product"], {"revenue": "sum", "order_id": "count"})
    """
    groups = ", ".join(f'"{c}"' for c in group_by)
    aggs = ", ".join(f'{fn}("{col}") AS "{fn}_{col}"' for col, fn in measures.items())
    return duckdb.sql(
        f"SELECT {groups}, {aggs} FROM '{parquet_path}' GROUP BY {groups} ORDER BY {groups}"
    )


def get_rows(parquet_path, key_column, names, columns=None):
    """
    Pull the FULL text of a few specific rows - the two-pass lookup.
    Reads only the matching rows off disk.
    """
    picked = ", ".join(f'"{c}"' for c in columns) if columns else "*"
    quoted = ", ".join("'" + str(n).replace("'", "''") + "'" for n in names)
    return duckdb.sql(
        f'SELECT {picked} FROM \'{parquet_path}\' WHERE "{key_column}" IN ({quoted})'
    ).fetchall()

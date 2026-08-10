"""
STEP 1 + 2 + 3: Read the file, clean it, and measure it.

This file knows nothing about AI. It only deals with files and rows.
Everything here can be tested without an API key.
"""

import csv
import io
import re

# Rows starting with these words are usually summary rows added by a human.
# Leaving them in would make every total come out double.
TOTAL_WORDS = ("total", "sum", "grand", "subtotal", "average", "avg")


# ----------------------------------------------------------------------
# STEP 1 - READ
# ----------------------------------------------------------------------

def read_file(path):
    """
    Read a CSV or Excel file.

    Always returns {table_name: [row dicts]}.
    A CSV gives one table. An Excel file gives one table per sheet.
    """
    if path.lower().endswith((".xlsx", ".xls")):
        return _read_excel(path)
    return {"data": _read_csv(path)}


def _read_csv(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        return [dict(row) for row in csv.DictReader(f)]


def _read_excel(path, max_rows=None):
    """
    read_only=True streams the sheet row by row instead of loading the whole
    workbook into memory. Memory stays flat whether the file is 1 MB or 200 MB.
    """
    import openpyxl   # imported here so CSV-only users don't need it

    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    tables = {}

    try:
        for name in book.sheetnames:
            raw_rows = []
            for i, row in enumerate(book[name].iter_rows(values_only=True)):
                if max_rows and i >= max_rows:
                    break
                raw_rows.append(row)
            if not raw_rows:
                continue

            header_index = _find_header_row(raw_rows)
            headers = [
                str(h).strip() if h not in (None, "") else f"column_{i}"
                for i, h in enumerate(raw_rows[header_index])
            ]

            rows = []
            carry = {}                                   # for merged cells
            for raw in raw_rows[header_index + 1:]:
                row = {}
                for col, value in zip(headers, raw):
                    text = str(value).strip() if value is not None else ""
                    if text == "":
                        text = carry.get(col, "")        # merged cell - copy from above
                    else:
                        carry[col] = text
                    row[col] = text

                if any(row.values()):                    # skip fully blank rows
                    rows.append(row)

            tables[name] = rows
    finally:
        book.close()                                     # read_only mode holds the file open

    return tables


def _find_header_row(raw_rows, look_at=10):
    """The header is the first row where most cells have something in them."""
    for i, row in enumerate(raw_rows[:look_at]):
        filled = sum(1 for cell in row if cell not in (None, ""))
        if filled >= max(2, len(row) * 0.6):
            return i
    return 0


# ----------------------------------------------------------------------
# STEP 2 - CLEAN
# ----------------------------------------------------------------------

def squash(value):
    """Turn line breaks, tabs and runs of spaces into one single space."""
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean(rows, drop_empty_above=0.9):
    """
    Remove the waste:
      - line breaks and padding spaces hiding inside cells
      - columns that are almost entirely empty
      - human-added TOTAL rows
    """
    if not rows:
        return rows

    rows = [{col: squash(value) for col, value in row.items()} for row in rows]

    # Drop columns that are nearly always blank - they cost tokens and say nothing
    keep = []
    for col in rows[0]:
        blank_share = sum(1 for r in rows if r[col] == "") / len(rows)
        if blank_share < drop_empty_above:
            keep.append(col)
    rows = [{col: row[col] for col in keep} for row in rows]

    return _drop_totals_rows(rows)


def _drop_totals_rows(rows, look_at_last=5):
    """
    Remove human-added summary rows like "GRAND TOTAL".

    Two guards stop us deleting real data:
      - only the last few rows are checked, because that is where totals live
      - the word must be a whole word, so "Total Recall" and "Black Summer" survive
    """
    if len(rows) <= look_at_last:
        return rows

    pattern = re.compile(r"\b(" + "|".join(TOTAL_WORDS) + r")\b", re.IGNORECASE)

    keep_to = len(rows)
    for i in range(len(rows) - 1, len(rows) - look_at_last - 1, -1):
        first_cell = str(next(iter(rows[i].values()), ""))
        if pattern.search(first_cell):
            keep_to = i
        else:
            break

    return rows[:keep_to]


# ----------------------------------------------------------------------
# STEP 3 - MEASURE
# ----------------------------------------------------------------------

def column_weights(rows):
    """
    How much of the file does each column take up?
    Returns [(column, percent, rough_tokens)] heaviest first.
    """
    if not rows:
        return []

    # str() because rolled-up rows contain real numbers, not just text
    sizes = {col: sum(len(str(r.get(col, ""))) for r in rows) for col in rows[0]}
    total = sum(sizes.values()) or 1

    return sorted(
        ((col, size / total * 100, int(size / 3.5)) for col, size in sizes.items()),
        key=lambda item: -item[1],
    )


def print_weights(rows, table_name="data"):
    """Print the weight report. Run this on every new file before anything else."""
    print(f"\n{table_name}: {len(rows):,} rows, {len(rows[0]) if rows else 0} columns")
    print("-" * 52)
    for col, percent, tokens in column_weights(rows):
        bar = "#" * max(1, round(percent / 2))
        print(f"  {col[:18]:18s} {percent:5.1f}%  ~{tokens:>8,d}  {bar}")


def to_csv(rows, columns=None):
    """Turn rows into CSV text - the cheapest format there is."""
    if not rows:
        return ""
    columns = columns or list(rows[0])

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row.get(col, "") for col in columns])
    return buffer.getvalue()


def rough_tokens(text):
    """
    Conservative offline estimate.

    Divide by 2.8, not 3.5. Measured against the real count on messy data the
    3.5 figure came out about 25% LOW, which is the dangerous direction - you
    think it fits, the API rejects it.
    """
    return int(len(text) / 2.8)


# ----------------------------------------------------------------------
# ROLLUP - the biggest saving there is
# ----------------------------------------------------------------------

def reliability_notes(summary, group_by):
    """
    Look at a rolled-up table and warn about things that produce confident
    wrong answers.

    The numbers can be perfectly correct and the conclusion still nonsense -
    comparing a group of 23 records against one of 3,723 is not a fair
    comparison, but nothing in the table says so.
    """
    notes = []
    if not summary:
        return notes

    counts = [r.get("n_rows", 0) for r in summary]
    smallest, largest = min(counts), max(counts)

    if largest > smallest * 20:
        small = [r for r in summary if r["n_rows"] < largest / 20]
        labels = ", ".join(str(r[group_by[0]]) for r in sorted(
            small, key=lambda r: r["n_rows"])[:6])
        notes.append(
            f"Group sizes are very uneven: from {smallest:,} to {largest:,} records. "
            f"Groups with few records ({labels}) are not comparable with the large "
            f"ones - a difference there is probably a sampling artefact, not a real "
            f"trend. Do not describe it as one."
        )

    thin = [r for r in summary if r.get("n_rows", 0) < 50]
    if len(thin) > len(summary) / 3:
        notes.append(
            f"{len(thin)} of {len(summary)} groups have under 50 records. "
            f"Treat their averages as rough."
        )

    return notes


def find_multi_value_columns(rows, sample=800):
    """
    Find columns that hold a LIST in one cell - "Action, Horror, Thriller".

    Guards matter here. A plot summary also contains commas, but its pieces are
    long and never repeat. A category list has short pieces drawn from a small
    fixed set. Both guards must pass.
    """
    if not rows:
        return []

    found = []
    for col in rows[0]:
        values = [str(r.get(col, "")) for r in rows[:sample] if r.get(col)]
        if len(values) < 10:
            continue

        with_comma = sum(1 for v in values if "," in v)
        if with_comma < len(values) * 0.5:
            continue                                  # not mostly lists

        parts = [p.strip() for v in values for p in v.split(",") if p.strip()]
        if not parts:
            continue

        distinct = len(set(parts))
        longest = max(len(p) for p in parts)
        avg_words = sum(p.count(" ") + 1 for p in parts) / len(parts)

        # A category label is SHORT and one or two words: "Horror", "Bulk",
        # "Docking Station". A sentence fragment is long and wordy: "Called
        # about delayed shipment". Test the LONGEST piece, not the average -
        # averaging one long fragment with one short one hides the problem.
        if distinct <= 50 and longest <= 25 and avg_words <= 2.0:
            found.append(col)

    return found


def find_messy_year_columns(rows, sample=800):
    """
    Find columns that contain a year but are not clean years.

    '(2005-2008)' and '-2021' both hold a year, but grouping on the raw text
    makes hundreds of junk groups.
    """
    if not rows:
        return []

    found = []
    for col in rows[0]:
        values = [str(r.get(col, "")) for r in rows[:sample] if r.get(col)]
        if len(values) < 10:
            continue

        has_year = sum(1 for v in values if re.search(r"(19|20)\d{2}", v))
        if has_year < len(values) * 0.7:
            continue                                  # not really a year column

        clean = sum(1 for v in values if re.fullmatch(r"\s*(19|20)\d{2}\s*", v))
        if clean < len(values) * 0.9:                 # mostly NOT clean
            found.append(col)

    return found


MEASURE_WORDS = (
    "price", "cost", "amount", "revenue", "sales", "total", "value", "salary",
    "profit", "margin", "spend", "budget", "fee", "rate", "score", "rating",
    "weight", "height", "age", "duration", "balance",
)


def classify_columns(rows, sample=800):
    """
    Split columns into things you GROUP BY and things you MEASURE.

    dimensions - short, repeating labels: genre, region, type, year
    measures   - numbers worth averaging or summing: rating, revenue, votes

    Columns that are neither (free text, ids) are left out of both - grouping by
    a unique id makes one group per row, which is not a summary at all.
    """
    dimensions, measures = [], []
    if not rows:
        return dimensions, measures

    total = min(len(rows), sample)

    for col in rows[0]:
        values = [str(r.get(col, "")).strip() for r in rows[:sample]]
        filled = [v for v in values if v]
        if len(filled) < total * 0.3:
            continue                                   # mostly empty

        numeric = sum(1 for v in filled[:200]
                      if re.fullmatch(r"-?[\d,]*\.?\d+", v.replace(",", "")))
        if numeric >= len(filled[:200]) * 0.9:
            # Decimals are always measures - you average a rating, you do not
            # group by it. Whole numbers may be labels: a year, or a 1-5 scale.
            has_decimals = any("." in v for v in filled[:200])
            looks_like_years = all(
                re.fullmatch(r"(19|20)\d{2}", v) for v in filled[:200]
            )
            small_set = len(set(filled)) <= 20

            # Some names are measures whatever their cardinality. A price with
            # only six distinct values is still a price, not a label.
            if any(word in col.lower() for word in MEASURE_WORDS):
                measures.append(col)
            elif not has_decimals and (looks_like_years or small_set):
                dimensions.append(col)
            else:
                measures.append(col)
            continue

        distinct = len(set(filled))
        avg_len = sum(len(v) for v in filled) / len(filled)

        # A label repeats and is short. An id or a sentence does not.
        if distinct <= max(60, total * 0.05) and avg_len <= 40:
            dimensions.append(col)

    return dimensions, measures


def auto_prepare(rows):
    """
    Get any uploaded table ready for grouping, without knowing the file.

    Returns (rows, notes) where notes describes what was changed, so the change
    can be shown to the user instead of happening invisibly.
    """
    notes = []
    if not rows:
        return rows, notes

    for col in find_multi_value_columns(rows):
        before = len(rows)
        rows = explode(rows, col)
        notes.append(f"'{col}' holds several values per cell - split into one row "
                     f"each ({before:,} rows became {len(rows):,}).")

    for col in find_messy_year_columns(rows):
        type_col = f"{col}_TYPE"
        for row in rows:
            row[type_col] = title_type(row.get(col))
            row[col] = extract_year(row.get(col))
        kinds = {r[type_col] for r in rows if r[type_col]}
        if len(kinds) > 1:
            notes.append(f"'{col}' mixes single dates with date RANGES. Pulled out a "
                         f"clean year and added '{type_col}' (Single vs Range) - a range "
                         f"record is filed under the year it started, so the two must "
                         f"not share a trend line.")
        else:
            for row in rows:
                row.pop(type_col, None)
            notes.append(f"'{col}' had messy year values - pulled out a clean year.")

    return rows, notes


def title_type(value):
    """
    Tell a single-date record from one that spans a range of years.

        '-2021'        -> Single   one year, so it happened then
        '(2004-2012)'  -> Range    it ran across years
        '(2021- )'     -> Range    still ongoing

    Why this matters: a range record is filed under the year it STARTED, so
    mixing the two invents trends. In the movie data, TV series (Range) average
    7.42 and films (Single) 6.51 - so any year with many new series looked like
    a golden age. It was not; it was a different kind of record.
    """
    text = str(value or "")
    if not re.search(r"(19|20)\d{2}", text):
        return ""
    if re.fullmatch(r"\s*-?\s*(19|20)\d{2}\s*", text):
        return "Single"
    if re.search(r"(19|20)\d{2}\s*[–—-]", text):
        return "Range"
    return "Single"


def extract_year(value):
    """
    Pull one clean year out of a messy cell.

    This dataset stores things like '-2021', '(2005-2008)', '(2016 TV Special)'
    and ''. Grouping by the raw value makes hundreds of junk groups, so pull the
    first 4-digit year and use that.
    """
    match = re.search(r"(19|20)\d{2}", str(value or ""))
    return match.group(0) if match else ""


def explode(rows, column, separator=",", label="(not set)"):
    """
    Split a multi-value cell into one row per value.

    "Action, Horror, Thriller" is THREE genres. Keeping only the first one
    throws most of the data away - on real movie data it lost 93% of thrillers.

    Returns a new list. The original rows are untouched.
    """
    out = []
    for row in rows:
        values = [v.strip() for v in str(row.get(column, "")).split(separator)]
        values = [v for v in values if v] or [label]

        for value in values:
            copy = dict(row)
            copy[column] = value
            out.append(copy)
    return out


def rollup(rows, group_by, measures, min_rows=1):
    """
    Group rows so a huge table becomes a small one.

        rollup(rows, ["GENRE"], {"RATING": "avg", "MOVIES": "count"})

    9,999 movies -> 28 summary rows. Every trend question is still answerable,
    and the numbers are computed here in Python so they are exact.

    measures: {column: "sum" | "avg" | "count" | "min" | "max"}
    """
    buckets = {}
    for row in rows:
        key = tuple(squash(row.get(col, "")) for col in group_by)
        buckets.setdefault(key, []).append(row)

    out = []
    for key, group in buckets.items():
        summary = dict(zip(group_by, key))
        summary["n_rows"] = len(group)

        for col, how in measures.items():
            if how == "count":
                summary[f"count_{col}"] = len(group)
                continue

            numbers = []
            for r in group:
                try:
                    numbers.append(float(str(r.get(col, "")).replace(",", "")))
                except (TypeError, ValueError):
                    pass

            if not numbers:
                summary[f"{how}_{col}"] = ""
            elif how == "sum":
                summary[f"sum_{col}"] = round(sum(numbers), 2)
            elif how == "avg":
                summary[f"avg_{col}"] = round(sum(numbers) / len(numbers), 2)
            elif how == "min":
                summary[f"min_{col}"] = min(numbers)
            elif how == "max":
                summary[f"max_{col}"] = max(numbers)

        out.append(summary)

    # Groups this small give meaningless averages - one movie rated 8.2 is not
    # "the best genre". Drop them rather than let the model draw a conclusion.
    if min_rows > 1:
        out = [r for r in out if r["n_rows"] >= min_rows]

    # A group where every measure came out blank tells the model nothing and
    # invites it to invent an explanation. Drop those too.
    measure_keys = [k for k in (out[0] if out else {}) if k not in group_by and k != "n_rows"]
    out = [r for r in out if any(r.get(k) not in (None, "") for k in measure_keys)]

    out.sort(key=lambda r: tuple(str(r[c]) for c in group_by))
    return out

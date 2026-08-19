"""
Tests for agent/workbook.py — reading a whole workbook as structure.

The generality concern here is that a workbook is not a CSV with tabs: sheet
names carry spaces and non-Latin characters, sheets are empty or header-only,
formulas reference other sheets in several notations, and a real file may have
dozens of tabs and hundreds of columns. Each of those is exercised below on a
workbook built for the purpose, and the card is checked for the property that
makes it usable at all — that its size does not grow with the DATA, only with
the number of sheets and columns.

Run via:  python tests/run_tests.py
"""

import os

import pandas as pd

from agent.ml.data_pipeline import WORKSPACE_DIR, get_data_pipeline
from agent.workbook import (
    _functions_used,
    _referenced_sheets,
    _sanitise,
    inspect_workbook,
    load_workbook,
)

_TMP = "test_wb_tmp.xlsx"


def _build(sheets: dict, formulas: dict = None, path: str = _TMP) -> str:
    """sheets: {name: [[row], ...]}  formulas: {sheet: {coord: '=...'}}"""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    for name, cells in (formulas or {}).items():
        ws = wb[name]
        for coord, text in cells.items():
            ws[coord] = text
    wb.save(os.path.join(WORKSPACE_DIR, path))
    return path


def _cleanup(path: str = _TMP, *names: str) -> None:
    try:
        os.remove(os.path.join(WORKSPACE_DIR, path))
    except OSError:
        pass
    pipe = get_data_pipeline()
    for n in names:
        pipe.datasets.pop(n, None)
        pipe.sources.pop(n, None)


# ─────────────────────────── formula reference parsing ────────────────────────
# A missed reference silently drops an edge from the dependency graph, so every
# notation Excel writes has to be recognised.


def test_a_plain_cell_reference_is_found():
    assert _referenced_sheets("=Orders!B2") == {"Orders"}


def test_a_range_reference_is_found():
    assert _referenced_sheets("=SUM(Orders!B2:B99)") == {"Orders"}


def test_a_whole_column_reference_is_found():
    # VLOOKUP(...,Rates!A:B,...) and SUMIF(Orders!C:C,...) have NO row number.
    assert _referenced_sheets("=VLOOKUP(A2,Rates!A:B,2,FALSE)") == {"Rates"}
    assert _referenced_sheets("=SUMIF(Orders!C:C,A2,Orders!D:D)") == {"Orders"}


def test_a_whole_row_reference_is_found():
    assert _referenced_sheets("=SUM(Totals!1:1)") == {"Totals"}


def test_an_absolute_reference_is_found():
    assert _referenced_sheets("=C2*Rates!$B$4") == {"Rates"}


def test_a_quoted_sheet_name_with_spaces_is_found():
    assert _referenced_sheets("=SUM('Q1 Sales'!B2:B9)") == {"Q1 Sales"}


def test_several_sheets_in_one_formula_are_all_found():
    got = _referenced_sheets("=Orders!B2*Rates!A:B+'Old Data'!C1")
    assert got == {"Orders", "Rates", "Old Data"}


def test_a_formula_with_no_sheet_reference_yields_nothing():
    assert _referenced_sheets("=SUM(B2:B9)") == set()


def test_functions_are_extracted():
    assert {"SUM", "VLOOKUP"} <= _functions_used("=SUM(A1:A9)+VLOOKUP(B1,X!A:B,2,FALSE)")


# ─────────────────────────── name sanitising ──────────────────────────────────
# Registered names are used as Python identifiers by the sandbox bridge.


def test_sheet_names_become_usable_identifiers():
    for raw in ("Q1 Sales", "Q1-Sales (final)", "2023 Data", "売上", "a/b\\c"):
        assert _sanitise(raw).isidentifier(), raw


def test_an_empty_sheet_name_still_yields_something():
    assert _sanitise("").isidentifier()


# ─────────────────────────── the structural card ──────────────────────────────


def test_every_sheet_appears_in_the_card():
    _build({"Alpha": [["a"], [1]], "Beta": [["b"], [2]], "Gamma": [["c"], [3]]})
    out = inspect_workbook(_TMP)
    for name in ("Alpha", "Beta", "Gamma"):
        assert name in out, out
    _cleanup()


def test_inspect_loads_nothing_into_the_registry():
    _build({"Alpha": [["a"], [1]]})
    before = set(get_data_pipeline().datasets)
    inspect_workbook(_TMP)
    assert set(get_data_pipeline().datasets) == before
    _cleanup()


def test_cross_sheet_dependencies_are_reported():
    _build({"Src": [["v"], [1], [2]], "Calc": [["total"]]},
           formulas={"Calc": {"A2": "=SUM(Src!A2:A3)"}})
    out = inspect_workbook(_TMP)
    assert "'Calc'  ←  Src" in out, out
    _cleanup()


def test_a_dependency_through_a_whole_column_reference_is_reported():
    # The regression this pins: Rates!A:B was invisible to a cell-only pattern,
    # so a sheet could be described as independent when it was not.
    _build({"Rates": [["k", "v"], ["x", 1]], "Calc": [["out"]]},
           formulas={"Calc": {"A2": "=VLOOKUP(\"x\",Rates!A:B,2,FALSE)"}})
    assert "'Calc'  ←  Rates" in inspect_workbook(_TMP)
    _cleanup()


def test_a_workbook_with_no_formulas_says_so():
    _build({"One": [["a"], [1]], "Two": [["b"], [2]]})
    assert "no sheet references another" in inspect_workbook(_TMP)
    _cleanup()


def test_a_computed_column_is_not_described_as_empty():
    # openpyxl returns the CACHED value, and a program-written file has none.
    # Reporting those columns as empty would be a false statement about the data.
    _build({"Src": [["v"], [1], [2]], "Calc": [["region", "total"], ["north"]]},
           formulas={"Calc": {"B2": "=SUM(Src!A2:A3)"}})
    out = inspect_workbook(_TMP)
    assert "total(formula)" in out, out
    assert "total(empty)" not in out
    assert "never been evaluated" in out
    _cleanup()


def test_shared_columns_are_offered_as_join_keys():
    _build({"L": [["Product_ID", "x"], [1, "a"]], "R": [["Product_ID", "y"], [1, "b"]]})
    out = inspect_workbook(_TMP)
    assert "SHARED COLUMNS" in out and "product_id" in out
    _cleanup()


def test_the_card_does_not_grow_with_the_number_of_rows():
    # This is the property that makes a large workbook tractable at all.
    small = _build({"S": [["a", "b"]] + [[i, i * 2] for i in range(20)]}, path="wb_small.xlsx")
    big = _build({"S": [["a", "b"]] + [[i, i * 2] for i in range(20_000)]}, path="wb_big.xlsx")
    a, b = inspect_workbook(small), inspect_workbook(big)
    assert "20,000 rows" in b
    assert abs(len(a) - len(b)) < 200, (len(a), len(b))
    _cleanup(small); _cleanup(big)


def test_no_data_row_beyond_the_sample_reaches_the_card():
    rows = [["v"]] + [[f"SECRET{i}"] for i in range(500)]
    _build({"S": rows})
    out = inspect_workbook(_TMP)
    assert "SECRET0" in out                       # a few samples are intended
    assert "SECRET499" not in out                 # the bulk is not
    assert out.count("SECRET") <= 5
    _cleanup()


def test_a_wide_sheet_is_truncated_rather_than_dumped():
    _build({"W": [[f"c{i}" for i in range(300)], list(range(300))]})
    out = inspect_workbook(_TMP)
    assert "more" in out and "c299" not in out
    _cleanup()


def test_an_empty_sheet_does_not_break_the_scan():
    _build({"Empty": [], "Real": [["a"], [1]]})
    out = inspect_workbook(_TMP)
    assert "Empty" in out and "Real" in out
    _cleanup()


def test_a_header_only_sheet_reports_zero_rows():
    _build({"H": [["a", "b"]]})
    assert "0 rows" in inspect_workbook(_TMP)
    _cleanup()


def test_a_missing_file_is_an_error_not_a_crash():
    assert "not found" in inspect_workbook("no_such_workbook.xlsx")


def test_a_path_outside_the_workspace_is_refused():
    assert "escapes the workspace" in inspect_workbook("../../../etc/passwd")


# ─────────────────────────── loading ──────────────────────────────────────────


def test_every_sheet_is_registered_under_its_own_name():
    _build({"Alpha": [["a"], [1]], "Beta": [["b"], [2]]})
    out = load_workbook(_TMP, prefix="wbtest")
    pipe = get_data_pipeline()
    assert "wbtest_Alpha" in pipe.datasets and "wbtest_Beta" in pipe.datasets
    assert "registered as 'wbtest_Alpha'" in out
    _cleanup(_TMP, "wbtest_Alpha", "wbtest_Beta")


def test_only_the_named_sheets_are_loaded():
    _build({"Alpha": [["a"], [1]], "Beta": [["b"], [2]]})
    load_workbook(_TMP, prefix="wbsel", sheets=["Alpha"])
    pipe = get_data_pipeline()
    assert "wbsel_Alpha" in pipe.datasets
    assert "wbsel_Beta" not in pipe.datasets
    _cleanup(_TMP, "wbsel_Alpha")


def test_an_unknown_sheet_name_is_reported_not_ignored():
    _build({"Alpha": [["a"], [1]]})
    out = load_workbook(_TMP, prefix="wbunk", sheets=["Alpha", "Nope"])
    assert "Nope" in out
    _cleanup(_TMP, "wbunk_Alpha", "wbunk")


def test_awkward_sheet_names_still_register_usably():
    _build({"Q1 Sales (final)": [["a"], [1]], "2023 数据": [["b"], [2]]})
    load_workbook(_TMP, prefix="wbodd")
    pipe = get_data_pipeline()
    made = [n for n in pipe.datasets if n.startswith("wbodd")]
    assert len(made) == 2, made
    for name in made:
        assert name.isidentifier(), name
    _cleanup(_TMP, *made)


def test_loaded_sheet_values_match_the_file():
    _build({"S": [["a", "b"], [1, "x"], [2, "y"]]})
    load_workbook(_TMP, prefix="wbval")
    df = get_data_pipeline().datasets["wbval"]
    assert list(df.columns) == ["a", "b"]
    assert df["a"].tolist() == [1, 2]
    _cleanup(_TMP, "wbval")


def test_a_single_sheet_workbook_keeps_the_plain_prefix():
    _build({"Only": [["a"], [1]]})
    load_workbook(_TMP, prefix="wbone")
    assert "wbone" in get_data_pipeline().datasets
    _cleanup(_TMP, "wbone")


def test_loaded_sheets_are_reachable_from_the_sandbox():
    from agent.data_bridge import reset, run_with_datasets
    reset()
    _build({"Nums": [["v"], [10], [20], [30]]})
    load_workbook(_TMP, prefix="wbsand")
    out = run_with_datasets("print('TOTAL', int(wbsand['v'].sum()))")
    assert "TOTAL 60" in out, out
    _cleanup(_TMP, "wbsand"); reset()

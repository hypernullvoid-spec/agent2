"""
Whole-workbook ingestion — read every tab at once, and hand the model a
structural picture rather than the rows.

Why this exists
───────────────
load_excel takes ONE sheet at a time and, faced with a multi-sheet file, refuses
and returns a list of names. For a workbook whose meaning lives in the relations
BETWEEN tabs — a summary sheet fed by an orders sheet fed by a rates sheet —
that forces the model to load blind, one sheet per call, with no way to see that
the tabs are connected at all.

The approach here is the one that makes large workbooks tractable: never put the
data in the context, put the STRUCTURE there. A card describing sheets, shapes,
column types, a couple of sample rows, the formulas and what they reference is a
few hundred tokens for a workbook of any size, and it is enough for the model to
decide which tab to compute over. The rows themselves stay in the registry and
are reached from run_python via agent/data_bridge.py.

Formulas
────────
pandas returns a formula cell's CACHED VALUE and discards the formula, so a
column computed as `=Orders!D2*Rates!B$4` is indistinguishable from one someone
typed by hand. That distinction decides whether a number is evidence or a
restatement of other numbers, so the formulas are read separately with openpyxl
(data_only=False) and reported alongside the values.

Everything is read in openpyxl's read_only streaming mode, so a workbook far
larger than memory can still be described.
"""

import os
import re
from typing import Optional

MAX_SHEETS_SHOWN = 40
MAX_COLS_SHOWN = 25
MAX_SAMPLE_ROWS = 3
MAX_FORMULAS_SHOWN = 25
SAMPLE_SCAN_ROWS = 200          # rows read to infer a column's type
CELL_TEXT_WIDTH = 18


def _safe_path(path: str) -> str:
    from agent.ml.data_pipeline import WORKSPACE_DIR
    full = os.path.abspath(os.path.join(WORKSPACE_DIR, path))
    if not (full == WORKSPACE_DIR or full.startswith(WORKSPACE_DIR + os.sep)):
        raise ValueError(f"Path '{path}' escapes the workspace directory")
    return full


def _sanitise(name: str) -> str:
    """A sheet name is free text ('Q1 Sales (final)'); a dataset name is used as
    a Python identifier by the sandbox bridge, so it has to be reduced to one."""
    cleaned = re.sub(r"\W+", "_", str(name)).strip("_")
    if not cleaned:
        cleaned = "sheet"
    if cleaned[0].isdigit():
        cleaned = f"s_{cleaned}"
    return cleaned


# ── formula parsing ───────────────────────────────────────────────────────────
# A cross-sheet reference comes in more shapes than a plain cell address, and
# missing one silently drops an edge from the dependency graph: VLOOKUP over
# Rates!A:B and SUMIF over Orders!C:C are WHOLE-COLUMN references with no row
# number at all, which a cell-only pattern never sees. All three forms —
# cell/range, whole column, whole row — are matched. The quoted spelling is what
# Excel writes whenever a sheet name contains a space.
_SHEET_REF = re.compile(
    r"""(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!            # Sheet! or 'My Sheet'!
        (?:\$?[A-Z]{1,3}\$?\d+                              # A1  /  $B$7
          |\$?[A-Z]{1,3}:\$?[A-Z]{1,3}                       # A:B  (whole columns)
          |\$?\d+:\$?\d+)                                   # 1:5  (whole rows)
    """, re.VERBOSE)
_FUNCS = re.compile(r"\b([A-Z][A-Z0-9.]{1,})\s*\(")


def _referenced_sheets(formula: str) -> set:
    out = set()
    for quoted, bare in _SHEET_REF.findall(formula or ""):
        out.add(quoted or bare)
    return out


def _functions_used(formula: str) -> set:
    return set(_FUNCS.findall((formula or "").upper()))


# ── type inference from a sample ──────────────────────────────────────────────

def _describe_value(value) -> str:
    import datetime as _dt
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return "date"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    text = str(value)
    if text.startswith("="):
        return "formula"
    return "text"


def _column_kind(values: list) -> str:
    kinds = [_describe_value(v) for v in values if v is not None]
    if not kinds:
        return "empty"
    ranked = ("formula", "date", "text", "float", "int", "bool")
    present = [k for k in ranked if k in kinds]
    if not present:
        return "text"
    if present[0] in ("float", "int") and {"float", "int"} <= set(kinds):
        return "float"
    return present[0]


def _clip(value, width: int = CELL_TEXT_WIDTH) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return text if len(text) <= width else text[: width - 1] + "…"


# ── the structural scan ───────────────────────────────────────────────────────

def _scan(path: str) -> dict:
    """Read structure only — never the whole sheet into memory.

    Returns {sheet: {rows, cols, headers, kinds, samples, formulas, refs, funcs}}.
    """
    from openpyxl import load_workbook as _open
    values_wb = _open(path, read_only=True, data_only=True)
    formula_wb = _open(path, read_only=True, data_only=False)
    try:
        report = {}
        for sheet_name in values_wb.sheetnames:
            vws = values_wb[sheet_name]
            headers, kinds, samples = [], [], []
            column_values: dict = {}
            for row_index, row in enumerate(vws.iter_rows(values_only=True)):
                if row_index == 0:
                    headers = [_clip(c, 40) or f"col{i + 1}"
                               for i, c in enumerate(row)]
                    continue
                if row_index <= MAX_SAMPLE_ROWS:
                    samples.append(row)
                for col_index, cell in enumerate(row):
                    column_values.setdefault(col_index, []).append(cell)
                if row_index >= SAMPLE_SCAN_ROWS:
                    break
            kinds = [_column_kind(column_values.get(i, [])) for i in range(len(headers))]

            fws = formula_wb[sheet_name]
            formulas, refs, funcs = [], set(), set()
            formula_cols: set = set()
            formula_at: dict = {}
            for row in fws.iter_rows():
                for cell in row:
                    text = cell.value
                    if isinstance(text, str) and text.startswith("="):
                        if len(formulas) < MAX_FORMULAS_SHOWN * 4:
                            formulas.append((cell.coordinate, text))
                        formula_cols.add(cell.column - 1)
                        if cell.row - 1 <= MAX_SAMPLE_ROWS:
                            formula_at[(cell.row - 1, cell.column - 1)] = text
                        refs |= _referenced_sheets(text)
                        funcs |= _functions_used(text)
            # openpyxl reads the CACHED value of a formula cell, and a workbook
            # written by a library rather than by Excel has no cached values at
            # all. Those cells come back as None, so a fully-computed column
            # reads as "empty" — which would tell the model the column holds no
            # data when it actually holds the sheet's entire point.
            kinds = ["formula" if i in formula_cols and k in ("empty", "blank") else k
                     for i, k in enumerate(kinds)]
            samples = [
                tuple(formula_at.get((r_i + 1, c_i), cell)
                      for c_i, cell in enumerate(sample))
                for r_i, sample in enumerate(samples)
            ]
            uncached = bool(formula_cols) and any(
                k == "formula" for k in kinds)
            report[sheet_name] = {
                "rows": max(0, (vws.max_row or 1) - 1),
                "cols": vws.max_column or len(headers),
                "headers": headers,
                "kinds": kinds,
                "samples": samples,
                "formulas": formulas,
                "refs": {r for r in refs if r != sheet_name},
                "funcs": funcs,
                "uncached": uncached,
            }
        return report
    finally:
        values_wb.close()
        formula_wb.close()


# ── the card ──────────────────────────────────────────────────────────────────

def _render(path: str, scan: dict, registered: Optional[dict] = None) -> str:
    lines = [f"WORKBOOK '{path}' — {len(scan)} sheet(s)", ""]
    shown = list(scan.items())[:MAX_SHEETS_SHOWN]

    lines.append("SHEETS")
    for name, info in shown:
        tag = ""
        if registered and name in registered:
            tag = f"   → registered as '{registered[name]}'"
        formula_note = f"   [{len(info['formulas'])} formula cell(s)]" if info["formulas"] else ""
        lines.append(f"  '{name}'  {info['rows']:,} rows × {info['cols']} cols{tag}{formula_note}")
        if info.get("uncached"):
            lines.append("      NOTE: the computed columns hold no stored value — this file "
                         "was written by a program, not saved by Excel, so the formulas "
                         "have never been evaluated. Reading this sheet gives blanks; "
                         "recompute from the source sheets instead.")
        cols = list(zip(info["headers"], info["kinds"]))[:MAX_COLS_SHOWN]
        if cols:
            lines.append("      " + "  ".join(f"{h}({k})" for h, k in cols)
                         + (f"  …+{len(info['headers']) - len(cols)} more"
                            if len(info["headers"]) > len(cols) else ""))
        for sample in info["samples"][:MAX_SAMPLE_ROWS]:
            lines.append("      e.g. " + " | ".join(_clip(c) for c in sample[:MAX_COLS_SHOWN]))
    if len(scan) > len(shown):
        lines.append(f"  …+{len(scan) - len(shown)} more sheet(s) not shown")

    # formulas — what the numbers are MADE OF
    with_formulas = [(n, i) for n, i in scan.items() if i["formulas"]]
    if with_formulas:
        lines += ["", "FORMULAS — these values are computed, not recorded"]
        for name, info in with_formulas:
            lines.append(f"  '{name}':")
            for coord, text in info["formulas"][:MAX_FORMULAS_SHOWN]:
                lines.append(f"    {coord}  {_clip(text, 90)}")
            if len(info["formulas"]) > MAX_FORMULAS_SHOWN:
                lines.append(f"    …+{len(info['formulas']) - MAX_FORMULAS_SHOWN} more formula cell(s)")
            if info["funcs"]:
                lines.append(f"    functions used: {', '.join(sorted(info['funcs'])[:12])}")

    # cross-sheet dependency graph
    deps = {n: i["refs"] for n, i in scan.items() if i["refs"]}
    lines += ["", "CROSS-SHEET DEPENDENCIES"]
    if deps:
        for name, refs in deps.items():
            lines.append(f"  '{name}'  ←  {', '.join(sorted(refs))}")
        sources = [n for n in scan if n not in deps]
        if sources:
            lines.append(f"  sheet(s) that depend on nothing else (raw input): "
                         f"{', '.join(sources)}")
        lines.append("  A sheet that derives from another is NOT independent evidence — "
                     "totals there restate the source, so do not present both as "
                     "separate confirmation of the same finding.")
    else:
        lines.append("  none — no sheet references another, so each stands on its own.")

    # shared column names → candidate joins
    by_column: dict = {}
    for name, info in scan.items():
        for header in info["headers"]:
            if header:
                by_column.setdefault(str(header).strip().lower(), set()).add(name)
    shared = {c: s for c, s in by_column.items() if len(s) > 1}
    if shared:
        lines += ["", "SHARED COLUMNS (candidate join keys)"]
        for col, sheets in sorted(shared.items())[:15]:
            lines.append(f"  '{col}': {', '.join(sorted(sheets))}")
    return "\n".join(lines)


# ── the two public operations ─────────────────────────────────────────────────

def inspect_workbook(path: str) -> str:
    """Describe every sheet without loading any of them into the registry."""
    try:
        full = _safe_path(path)
    except ValueError as e:
        return f"Error: {e}"
    if not os.path.exists(full):
        return f"Error: file not found: {path}"
    try:
        scan = _scan(full)
    except Exception as e:  # noqa: BLE001
        return f"Error reading workbook '{path}': {type(e).__name__}: {e}"
    if not scan:
        return f"'{path}' has no sheets."
    return (_render(path, scan)
            + "\n\nNothing was loaded. Call load_workbook to register the sheets you need.")


def load_workbook(path: str, prefix: Optional[str] = None,
                  sheets: Optional[list] = None) -> str:
    """Load every sheet (or the named ones) into the registry, one dataset each."""
    import pandas as pd
    from agent.ml.data_pipeline import get_data_pipeline
    try:
        full = _safe_path(path)
    except ValueError as e:
        return f"Error: {e}"
    if not os.path.exists(full):
        return f"Error: file not found: {path}"
    try:
        scan = _scan(full)
    except Exception as e:  # noqa: BLE001
        return f"Error reading workbook '{path}': {type(e).__name__}: {e}"
    if not scan:
        return f"'{path}' has no sheets."

    wanted = list(scan) if not sheets else [s for s in sheets if s in scan]
    unknown = [s for s in (sheets or []) if s not in scan]
    if not wanted:
        return (f"Error: none of {sheets} are sheets in '{path}'. "
                f"Available: {list(scan)}")

    base = _sanitise(prefix) if prefix else _sanitise(os.path.splitext(os.path.basename(path))[0])
    pipe = get_data_pipeline()
    registered, failed = {}, []
    for sheet in wanted:
        try:
            df = pd.read_excel(full, sheet_name=sheet)
        except Exception as e:  # noqa: BLE001 — one bad sheet must not sink the rest
            failed.append(f"{sheet} ({type(e).__name__}: {e})")
            continue
        # The suffix is dropped only for a workbook that HAS one sheet — not for
        # one sheet picked out of many, where losing which tab it came from is
        # exactly the information worth keeping.
        name = base if len(scan) == 1 else f"{base}_{_sanitise(sheet)}"
        pipe.datasets[name] = df
        pipe.sources[name] = f"excel:{path}#{sheet}"
        registered[sheet] = name

    out = [_render(path, {k: v for k, v in scan.items() if k in wanted or True},
                   registered=registered)]
    if failed:
        out.append("\nSHEETS THAT COULD NOT BE READ:\n  " + "\n  ".join(failed))
    if unknown:
        out.append(f"\nNot a sheet in this file (skipped): {unknown}")
    out.append(f"\nRegistered {len(registered)} sheet(s) into the dataset registry. "
               f"Use them BY NAME — in the analysis tools, or directly as DataFrame "
               f"variables inside run_python.")
    return "\n".join(out)


# ── swarn registration ────────────────────────────────────────────────────────

_SWARN_TOOLS = {
    "inspect_workbook": (
        "Describe EVERY sheet of an Excel workbook without loading any data: per-sheet "
        "shape, column names and types, sample rows, the formulas each sheet contains, "
        "which sheets feed which, and columns shared between sheets. Costs a few hundred "
        "tokens for a workbook of any size — use this FIRST on any .xlsx, before deciding "
        "what to load. Reading a multi-sheet workbook one sheet at a time hides the "
        "relationships between the tabs.",
        {"type": "object",
         "properties": {"path": {"type": "string",
                                 "description": "Relative path to the .xlsx inside the workspace."}},
         "required": ["path"]},
        inspect_workbook,
    ),
    "load_workbook": (
        "Load an Excel workbook's sheets into the dataset registry — ALL of them by "
        "default, one dataset per sheet, named '<file>_<sheet>'. Returns the same "
        "structural card as inspect_workbook, annotated with the name each sheet was "
        "registered under. Use this instead of calling load_excel once per sheet.",
        {"type": "object",
         "properties": {
             "path": {"type": "string", "description": "Relative path to the .xlsx inside the workspace."},
             "prefix": {"type": "string", "description": "Name prefix for the registered datasets (default: the file name)."},
             "sheets": {"type": "array", "items": {"type": "string"},
                        "description": "Only these sheets. Omit to load every sheet."},
         },
         "required": ["path"]},
        load_workbook,
    ),
}


def register_into_swarn() -> str:
    try:
        from agent.runtime.tools import TOOL_REGISTRY
    except Exception as e:  # noqa: BLE001
        return f"swarn registration skipped: {e}"
    count = 0
    for name, (desc, schema, fn) in _SWARN_TOOLS.items():
        if name in TOOL_REGISTRY:
            continue
        TOOL_REGISTRY[name] = {"description": desc, "schema": schema, "func": fn}
        count += 1
    return f"registered {count} workbook tool(s) into swarn"

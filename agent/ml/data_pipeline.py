"""
Phase 6: Data Ingestion & Validation

Gives the agent connectors for the common tabular data sources (CSV,
Excel, Parquet, SQL, cloud storage) and an automated validation layer
that profiles a loaded dataset and flags schema/null/outlier problems
*before* Phase 7's feature engineering or Phase 8's training ever sees it.

Design mirrors context_engine.py / sandbox.py: a singleton manager class
(`DataPipeline`) holds the in-memory dataset registry for the current
process, and tools.py exposes thin @tool wrappers around its methods.

Storage model
─────────────
Datasets are loaded into a small in-memory registry keyed by a name the
agent chooses (e.g. "train", "raw_sales"). Nothing is silently persisted
to disk — if the agent wants a cleaned copy saved, it calls
save_dataset() explicitly. This keeps the registry the single source of
truth within a session and avoids stale CSVs littering the workspace.

Validation philosophy
──────────────────────
validate_dataset() never raises and never blocks — it always returns a
structured report (schema, nulls, dtypes, outliers, duplicates) as text
the agent can read and act on. This is the same "errors as strings"
contract as tools.py, extended into "diagnostics as strings."

Phase boundary
───────────────
This module stops at "is the data sound and loaded into memory." Phase 7
(feature_engineering.py) consumes DataPipeline's registry to transform
columns; Phase 8 (model_training.py) consumes Phase 7's output to train.
None of these phases need to change agent_loop.py — same pattern as
Phase 2/3's sandbox.py and context_engine.py.
"""

import io
import os
from typing import Optional

import pandas as pd

from agent.paths import WORKSPACE_DIR, safe_path as _safe_path

MAX_PREVIEW_ROWS = 10
OUTLIER_Z_THRESHOLD = 3.0

# Two category columns overlapping is only a defect for one SHAPE of overlap.
# Tunable because "small minority" is a judgement call that differs by domain.
CATEGORY_SAME_DOMAIN_JACCARD = float(os.environ.get("SWARN_CATEGORY_SAME_DOMAIN", "0.6"))
CATEGORY_LEAK_MAX_SHARE = float(os.environ.get("SWARN_CATEGORY_LEAK_SHARE", "0.5"))

# ─── size guards ──────────────────────────────────────────────────────────────
#
# Both loaders used to hand the whole file to pandas and hope. Measured on a
# 7.6 MB / 200k-row .xlsx, pd.read_excel peaks at 12x the file size and the
# openpyxl object graph behind it at 62x — a Python object per cell, complete
# with its font and border. At 12x a 300 MB workbook wants ~3.6 GB, and the
# failure mode is a MemoryError with nothing the user can act on.
#
# The fix is not new: agent/workbook.py::_scan already reads workbooks
# read_only=True and streams rows with a cap, which measures at 2x. That
# technique simply never reached the loading path. It does now.
#
# A cap is only safe if it is LOUD. A silent truncation is worse than a crash:
# the crash tells you something went wrong, whereas half a file analysed
# confidently produces a report whose every total is wrong by an unknown
# amount. Everything read short is recorded and stated — see _register,
# note_truncation and data_report's limitations section.
MAX_LOAD_MB = float(os.environ.get("SWARN_MAX_LOAD_MB", "200"))
ROW_CAP = int(os.environ.get("SWARN_ROW_CAP", "2000000"))
STREAM_CHUNK_ROWS = int(os.environ.get("SWARN_STREAM_CHUNK", "100000"))

# Excel batches are deliberately far smaller than the CSV chunk size, and the
# difference is not cosmetic. pandas builds typed columns straight from a CSV
# chunk, but openpyxl hands back Python tuples of boxed values, which cost
# several times what the finished column does — so the batch has to be
# converted and released often. Measured on a 7.6 MB / 200k-row workbook:
#
#   pd.read_excel (the old path)          101 MB   13.3x file
#   streaming, 100,000-row batches         61 MB    8.0x
#   streaming,  10,000-row batches         27 MB    3.5x
#   iterating and keeping nothing at all   17 MB    2.2x   <- the floor
#
# 10,000 sits close to the floor; larger batches give most of the saving back.
EXCEL_BATCH_ROWS = int(os.environ.get("SWARN_EXCEL_BATCH", "10000"))

# Below this share of distinct values a text column is worth storing as a
# category — usually a 5-10x saving on region/status/city columns, and free.
CATEGORY_CONVERT_RATIO = float(os.environ.get("SWARN_CATEGORY_CONVERT_RATIO", "0.5"))


# _safe_path lives in agent/paths.py — imported above under its original name.


def _file_mb(path: str) -> float:
    try:
        return os.path.getsize(path) / 1e6
    except OSError:
        return 0.0


def _size_label(mb: float) -> str:
    """'340 MB' / '1.4 GB' / '0.4 MB' — never '0 MB', which reads as a bug."""
    if mb >= 1000:
        return f"{mb / 1000:,.1f} GB"
    return f"{mb:,.0f} MB" if mb >= 1 else f"{mb:.1f} MB"


def _shrink(df: pd.DataFrame) -> pd.DataFrame:
    """Reduce memory without changing any answer. Lossless only.

    Repeated text stored as a category is the same values with the duplicates
    removed — a sum, a ranking and a correlation all come out bit-identical, and
    on the columns that matter (region, status, city, customer) it is most of
    the saving available. Measured on a 200k-row frame: 9.1 MB -> 3.9 MB.

    float64 -> float32 was tried here and REMOVED. It saves a further 1.6 MB on
    that same frame, and it moves a revenue total of 200,408,457.53 to
    200,408,464.00 — an error of 6.47. Relatively that is nothing; practically
    it is a figure that no longer reconciles with the number finance already
    has, produced by a tool whose entire purpose is to not quietly hand back
    wrong ones. A saving is not worth a discrepancy nobody can explain.

    Integers are left alone for the same reason: downcasting an id column
    silently changes values that overflow.
    """
    for column in df.columns:
        series = df[column]
        # pandas 3 gives text columns a 'str' dtype rather than 'object', so
        # testing for object alone silently skips every text column on a modern
        # pandas — which is most of what there is to save. Both are checked.
        if not (series.dtype == object or pd.api.types.is_string_dtype(series)):
            continue
        non_null = int(series.notna().sum())
        if non_null and series.nunique(dropna=True) / non_null < CATEGORY_CONVERT_RATIO:
            try:
                df[column] = series.astype("category")
            except (TypeError, ValueError):
                pass                           # mixed types — leave it as it is
    return df


def _excel_dimensions(path: str) -> dict:
    """Rows and columns per sheet, read from the workbook's own metadata.

    read_only=True means openpyxl streams rather than building the cell graph,
    so this costs about the same as opening the zip — the whole point is to
    learn the size WITHOUT paying to load it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — .xls and friends have no openpyxl reader
        return {}
    try:
        return {sheet: (wb[sheet].max_row or 0, wb[sheet].max_column or 0)
                for sheet in wb.sheetnames}
    finally:
        wb.close()


def _excel_engine() -> Optional[str]:
    """Prefer calamine when it is installed — a Rust reader, far cheaper than
    openpyxl. Absent, pandas picks its own default and nothing changes."""
    try:
        import python_calamine  # noqa: F401
    except ImportError:
        return None
    return "calamine"


def _stream_excel(path: str, sheet, cap: int, usecols=None) -> tuple:
    """Read a sheet row by row instead of materialising it. Returns (df, rows_seen).

    `rows_seen` counts data rows encountered INCLUDING any past the cap, so the
    caller can say "read 2,000,000 of 5,300,000" rather than only knowing it
    stopped.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        name = sheet if isinstance(sheet, str) else wb.sheetnames[sheet or 0]
        if name not in wb.sheetnames:
            raise KeyError(f"sheet {name!r} not found; workbook has {wb.sheetnames}")
        ws = wb[name]
        headers: list = []
        keep: Optional[list] = None
        batch: list = []
        frames: list = []
        kept = 0
        rows_seen = 0

        def flush():
            # Converting each batch as it fills is the point of this function.
            # Holding every row as a Python tuple until the end costs MORE than
            # pd.read_excel does — a 6-field tuple plus its six boxed values is
            # far heavier than six slots in a typed column — so a version that
            # streamed the read but batched nothing measured no better than the
            # reader it replaced. The tuples have to be converted and released
            # as we go, which is what keeps peak memory near the size of the
            # finished frame rather than a multiple of it.
            nonlocal batch
            if batch:
                frames.append(pd.DataFrame(batch, columns=headers))
                batch = []

        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index == 0:
                headers = [str(c) if c is not None else f"col{i + 1}"
                           for i, c in enumerate(row)]
                if usecols:
                    wanted = {str(c) for c in usecols}
                    keep = [i for i, h in enumerate(headers) if h in wanted]
                    headers = [headers[i] for i in keep]
                continue
            rows_seen += 1
            if kept >= cap:
                continue                       # keep counting, stop collecting
            batch.append(tuple(row[i] for i in keep) if keep is not None else row)
            kept += 1
            if len(batch) >= EXCEL_BATCH_ROWS:
                flush()
        flush()
        if not frames:
            return pd.DataFrame(columns=headers), rows_seen
        df = (frames[0] if len(frames) == 1
              else pd.concat(frames, ignore_index=True, copy=False))
        return df, rows_seen
    finally:
        wb.close()


class DataPipeline:
    """
    Holds the in-memory dataset registry for the current process and
    implements every Phase 6 connector + the validation report.

    One instance per process (see get_data_pipeline() below), matching
    the Sandbox/ContextEngine singleton pattern already used in this
    codebase.
    """

    def __init__(self):
        self.datasets: dict[str, pd.DataFrame] = {}
        self.sources: dict[str, str] = {}      # name → where it was loaded from
        self.saved: dict[str, tuple] = {}      # name → (abs path, shape when written)
        # name → what was left unread. Empty for every fully-loaded dataset, so
        # its presence alone means "this frame is not the whole file".
        self.truncation: dict[str, dict] = {}
        self._cleaner_cache: dict[str, object] = {}

    # ───────────────────────────────────────────────── ingestion connectors

    def load_csv(self, path: str, name: str, **read_kwargs) -> str:
        full = _safe_path(path)
        if not os.path.exists(full):
            return f"Error: file not found: {path}"
        size_mb = _file_mb(full)
        truncation = None
        try:
            if size_mb <= MAX_LOAD_MB or "nrows" in read_kwargs:
                df = pd.read_csv(full, **read_kwargs)
            else:
                # Big file: accumulate chunks up to the cap, then keep reading
                # only to COUNT what is being left behind. Counting the tail is
                # cheap and it is the difference between "read 2,000,000 rows"
                # and "read 2,000,000 of 5,300,000 rows" — the second is a fact
                # a reader can act on, the first is not.
                frames, kept, total = [], 0, 0
                reader = pd.read_csv(full, chunksize=STREAM_CHUNK_ROWS, **read_kwargs)
                for chunk in reader:
                    total += len(chunk)
                    if kept < ROW_CAP:
                        room = ROW_CAP - kept
                        frames.append(chunk.iloc[:room] if len(chunk) > room else chunk)
                        kept += min(room, len(chunk))
                df = _shrink(pd.concat(frames, ignore_index=True) if frames
                             else pd.DataFrame())
                if total > kept:
                    truncation = {"rows_read": kept, "rows_total": total,
                                  "strategy": f"first {kept:,} rows of a "
                                              f"{_size_label(size_mb)} CSV"}
        except Exception as e:
            return f"Error reading CSV '{path}': {type(e).__name__}: {e}"
        return self._register(name, df, source=f"csv:{path}", truncation=truncation)

    def load_excel(self, path: str, name: str, sheet_name=0, **read_kwargs) -> str:
        full = _safe_path(path)
        if not os.path.exists(full):
            return f"Error: file not found: {path}"
        size_mb = _file_mb(full)
        truncation = None

        # A workbook big enough to matter is read by streaming rows rather than
        # by materialising the sheet — the 2x path instead of the 12x one. Small
        # files keep the pandas reader, which is faster there and handles the
        # multi-sheet dict return the agent relies on.
        heavy = size_mb > MAX_LOAD_MB and full.lower().endswith((".xlsx", ".xlsm"))
        if heavy and not isinstance(sheet_name, (list, type(None))):
            try:
                df, rows_seen = _stream_excel(full, sheet_name, ROW_CAP,
                                              read_kwargs.get("usecols"))
            except Exception as e:  # noqa: BLE001 — fall back rather than fail
                heavy = False
            else:
                df = _shrink(df)
                if rows_seen > len(df):
                    truncation = {"rows_read": len(df), "rows_total": rows_seen,
                                  "strategy": f"first {len(df):,} rows streamed from a "
                                              f"{_size_label(size_mb)} workbook"}
                return self._register(name, df, source=f"excel:{path}",
                                      truncation=truncation)

        engine = read_kwargs.pop("engine", None) or _excel_engine()
        try:
            df = pd.read_excel(full, sheet_name=sheet_name,
                               **({"engine": engine} if engine else {}), **read_kwargs)
        except Exception as e:
            return f"Error reading Excel '{path}': {type(e).__name__}: {e}"
        if isinstance(df, dict):
            sheets = ", ".join(df.keys())
            dims = _excel_dimensions(full)
            detail = "; ".join(f"{s} ({dims[s][0]:,}x{dims[s][1]})" for s in df if s in dims)
            return (
                f"'{path}' has multiple sheets: {sheets}. "
                + (f"Sizes: {detail}. " if detail else "")
                + "Pass sheet_name explicitly to load one."
            )
        return self._register(name, df, source=f"excel:{path}", truncation=truncation)

    def load_parquet(self, path: str, name: str) -> str:
        full = _safe_path(path)
        if not os.path.exists(full):
            return f"Error: file not found: {path}"
        try:
            df = pd.read_parquet(full)
        except Exception as e:
            return f"Error reading Parquet '{path}': {type(e).__name__}: {e}"
        return self._register(name, df, source=f"parquet:{path}")

    def load_sql(self, connection_string: str, query: str, name: str) -> str:
        try:
            from sqlalchemy import create_engine
        except ImportError:
            return "Error: SQL support requires 'pip install sqlalchemy'."
        try:
            engine = create_engine(connection_string)
            df = pd.read_sql(query, engine)
        except Exception as e:
            return f"Error querying database: {type(e).__name__}: {e}"
        return self._register(name, df, source="sql")

    def load_cloud_storage(self, uri: str, name: str) -> str:
        """
        Load a CSV/Parquet object from S3 (s3://) or GCS (gs://) directly
        into the registry. Requires boto3 (S3) or gcsfs (GCS) and valid
        credentials in the environment — this function does not manage
        credentials itself.
        """
        try:
            if uri.startswith("s3://"):
                if uri.endswith(".parquet"):
                    df = pd.read_parquet(uri)
                else:
                    df = pd.read_csv(uri)
            elif uri.startswith("gs://"):
                if uri.endswith(".parquet"):
                    df = pd.read_parquet(uri)
                else:
                    df = pd.read_csv(uri)
            else:
                return "Error: uri must start with 's3://' or 'gs://'."
        except Exception as e:
            return (
                f"Error reading '{uri}': {type(e).__name__}: {e}\n"
                "Check credentials (AWS env vars / GOOGLE_APPLICATION_CREDENTIALS) "
                "and that boto3/gcsfs is installed."
            )
        return self._register(name, df, source=f"cloud:{uri}")

    # ───────────────────────────────────────────────── validation

    @staticmethod
    def _overlapping_categories(df, max_levels: int = 60) -> list:
        """Two category columns sharing values usually means one leaked into the other.

        A real products file had 'Raksha Bandhan' — an OCCASION — sitting in the
        Category column alongside Sweets and Cake. Every check here passed it: the
        value is a non-null string, the schema infers cleanly, and no z-score
        applies to text. It was then reported as a legitimate 8.4%-of-revenue
        product category. Columns whose vocabularies overlap are worth a sentence,
        because nothing else in this report can see it.
        """
        candidates = []
        for col in df.columns:
            series = df[col]
            if not (series.dtype == object or str(series.dtype) in ("string", "str", "category")):
                continue
            # Single-character values count: 'M'/'F', grade bands and currency
            # codes are ordinary categoricals, and excluding them meant whole
            # columns were never checked at all.
            values = {str(v).strip() for v in series.dropna().unique()}
            values = {v for v in values if v}
            if 1 < len(values) <= max_levels:
                candidates.append((str(col), values))

        notes = []
        for i, (col_a, vals_a) in enumerate(candidates):
            for col_b, vals_b in candidates[i + 1:]:
                shared = vals_a & vals_b
                if not shared:
                    continue
                # WHICH KIND of overlap this is decides whether anything is wrong.
                # 'origin'/'destination' or 'billing_country'/'shipping_country'
                # share their WHOLE vocabulary by design — two roles of one entity
                # type, nothing to report. A stray value appearing in a column it
                # does not belong to shows up as a small minority of BOTH
                # vocabularies. Overlap shape separates them without needing to
                # know what the columns mean.
                jaccard = len(shared) / len(vals_a | vals_b)
                minority = (len(shared) / len(vals_a) < CATEGORY_LEAK_MAX_SHARE
                            and len(shared) / len(vals_b) < CATEGORY_LEAK_MAX_SHARE)
                sample = ", ".join(sorted(shared)[:4])
                more = f" (+{len(shared) - 4} more)" if len(shared) > 4 else ""

                if jaccard >= CATEGORY_SAME_DOMAIN_JACCARD:
                    # Same value domain. Only worth a word if the columns are also
                    # row-for-row identical, which makes one of them redundant.
                    try:
                        identical = bool(df[col_a].equals(df[col_b]))
                    except Exception:  # noqa: BLE001
                        identical = False
                    if identical:
                        notes.append(f"  '{col_a}' and '{col_b}' are identical row for row — "
                                     f"one is a redundant copy, not a second dimension.")
                    continue
                if minority:
                    notes.append(f"  '{col_a}' and '{col_b}' share {len(shared)} value(s): "
                                 f"{sample}{more}. That is a small minority of both columns, which "
                                 f"usually means a value belonging to one has leaked into the "
                                 f"other — verify it is a real '{col_a}' before reporting it as "
                                 f"one. (If these columns legitimately draw on the same list, "
                                 f"ignore this.)")
        return notes

    def validate_dataset(self, name: str) -> str:
        """
        Profile a loaded dataset and return a structured diagnostic report:
        shape, dtypes, null counts, duplicate rows, and per-numeric-column
        outliers (|z-score| > 3). Tries pandera first for schema-level
        checks (inferred-then-validated) and falls back to a hand-rolled
        profile if pandera isn't installed.
        """
        df = self.datasets.get(name)
        if df is None:
            return self._unknown_dataset_error(name)

        lines = [f"Validation report for '{name}'  (shape: {df.shape[0]} rows × {df.shape[1]} cols)"]

        # ── dtypes ──────────────────────────────────────────────
        lines.append("\ndtypes:")
        for col, dt in df.dtypes.items():
            lines.append(f"  {col}: {dt}")

        # ── nulls ───────────────────────────────────────────────
        null_counts = df.isnull().sum()
        nulls_present = null_counts[null_counts > 0]
        if len(nulls_present):
            lines.append("\nnulls:")
            for col, n in nulls_present.items():
                pct = 100 * n / len(df)
                lines.append(f"  {col}: {n} ({pct:.1f}%)")
        else:
            lines.append("\nnulls: none")

        # ── duplicates ──────────────────────────────────────────
        dup_count = int(df.duplicated().sum())
        lines.append(f"\nduplicate rows: {dup_count}")

        # ── outliers (numeric columns only, z-score method) ─────
        numeric_cols = df.select_dtypes(include="number").columns
        outlier_report = []
        for col in numeric_cols:
            series = df[col].dropna()
            if len(series) < 8 or series.std(ddof=0) == 0:
                continue
            z = (series - series.mean()) / series.std(ddof=0)
            n_outliers = int((z.abs() > OUTLIER_Z_THRESHOLD).sum())
            if n_outliers:
                outlier_report.append(f"  {col}: {n_outliers} values with |z| > {OUTLIER_Z_THRESHOLD}")
        lines.append("\noutliers (z-score method, numeric columns):")
        lines.extend(outlier_report if outlier_report else ["  none detected"])

        # ── categorical consistency ─────────────────────────────
        overlaps = self._overlapping_categories(df)
        lines.append("\ncategory consistency:")
        lines.extend(overlaps if overlaps else
                     ["  nothing to flag — no value looks out of place in its column"])

        # ── pandera schema check, if available ───────────────────
        try:
            import pandera as pa
            schema = pa.infer_schema(df)
            try:
                schema.validate(df, lazy=True)
                lines.append("\npandera inferred-schema check: passed")
            except pa.errors.SchemaErrors as se:
                lines.append("\npandera inferred-schema check: violations found")
                lines.append(str(se.failure_cases.head(20)))
        except ImportError:
            lines.append(
                "\n(pandera not installed — install_package('pandera') for "
                "stricter schema validation; profile above used pandas only)"
            )

        lines.append(
            f"\nNext step suggestion: if nulls/outliers/duplicates look "
            f"problematic, handle them before calling profile_dataset/engineer_features "
            f"on '{name}'."
        )
        return "\n".join(lines)

    # ───────────────────────────────────────────────── inspection helpers

    def preview_dataset(self, name: str, n: int = MAX_PREVIEW_ROWS) -> str:
        df = self.datasets.get(name)
        if df is None:
            return self._unknown_dataset_error(name)
        buf = io.StringIO()
        df.head(n).to_string(buf)
        return (
            f"'{name}'  shape={df.shape[0]}×{df.shape[1]}\n\n"
            f"{buf.getvalue()}"
        )

    def list_datasets(self) -> str:
        if not self.datasets:
            return "No datasets loaded. Use load_csv / load_excel / load_parquet / load_sql first."
        lines = ["Loaded datasets:"]
        for name, df in self.datasets.items():
            lines.append(f"  {name}: {df.shape[0]} rows × {df.shape[1]} cols")
        return "\n".join(lines)

    def save_dataset(self, name: str, path: str) -> str:
        df = self.datasets.get(name)
        if df is None:
            return self._unknown_dataset_error(name)
        full = _safe_path(path)
        os.makedirs(os.path.dirname(full), exist_ok=True)
        try:
            if path.endswith(".parquet"):
                df.to_parquet(full, index=False)
            else:
                df.to_csv(full, index=False)
        except Exception as e:
            return f"Error saving dataset: {type(e).__name__}: {e}"
        # the file now matches the registry copy, so reading it back is safe —
        # record that so tooling does not warn about a correct action
        self.saved[name] = (full, df.shape)
        return f"Saved '{name}' ({df.shape[0]} rows) to {path}"

    # ───────────────────────────────────────────────── internals

    def _register(self, name: str, df: pd.DataFrame, source: str,
                  truncation: Optional[dict] = None) -> str:
        self.datasets[name] = df
        # remembering the origin lets tooling warn when code re-reads the raw
        # file after a cleaned version already exists in the registry
        self.sources[name] = source
        if truncation:
            self.truncation[name] = dict(truncation)
            # Recorded in the ledger as well as the registry: the registry is
            # per-process working state, the ledger is what the report is built
            # from. A partial read has to reach the document, not just the agent.
            try:
                from agent.data_analysis import note_truncation
                note_truncation(name, truncation)
            except Exception:  # noqa: BLE001 — recording must not break a load
                pass
        else:
            self.truncation.pop(name, None)

        lines = [
            f"Loaded '{name}' from {source}  "
            f"({df.shape[0]} rows × {df.shape[1]} cols)",
            f"Columns: {', '.join(str(c) for c in df.columns)}",
        ]
        if truncation:
            read, total = truncation["rows_read"], truncation["rows_total"]
            lines.append(
                f"WARNING — THIS IS NOT THE WHOLE FILE. Read {read:,} of {total:,} rows "
                f"({read / total:.1%}); {total - read:,} row(s) were not loaded "
                f"({truncation['strategy']}). Every total, count and average you compute "
                f"from '{name}' describes only the part that was read, and the rows left out "
                f"are the LAST ones in the file — if it is sorted by date, a whole period is "
                f"missing. Say so whenever you quote a figure. To read more, raise "
                f"SWARN_ROW_CAP, or pass usecols to load fewer columns and fit more rows."
            )
        lines.append(f"Run validate_dataset('{name}') before using it downstream.")
        return "\n".join(lines)

    def _unknown_dataset_error(self, name: str) -> str:
        known = ", ".join(self.datasets.keys()) or "(none loaded)"
        return f"Error: no dataset named '{name}' is loaded. Loaded datasets: {known}"


# ─── singleton, matching sandbox.py / context_engine.py ───────────────────────

_pipeline: Optional[DataPipeline] = None


def get_data_pipeline() -> DataPipeline:
    global _pipeline
    if _pipeline is None:
        _pipeline = DataPipeline()
    return _pipeline
